"""
╔══════════════════════════════════════════════════════════════════╗
║        QR Certificate Generator Pro v2.0 — Streamlit App        ║
║        Developed By: Abdul Samad | SBBU NAWABSHAH               ║
╚══════════════════════════════════════════════════════════════════╝

INSTALL:
    pip install streamlit pillow qrcode[pil] reportlab openpyxl

RUN:
    streamlit run app.py

DEPLOY FREE:
    GitHub → streamlit.io/cloud → done!
"""

import streamlit as st
from PIL import Image, ImageDraw, ImageFont
import qrcode
import io
import json
import zipfile
from datetime import datetime, date
from reportlab.pdfgen import canvas as pdf_canvas
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.utils import ImageReader
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────────
#  Page Config  (MUST be first Streamlit call)
# ──────────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="QR Certificate Generator Pro",
    page_icon="🎓",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ──────────────────────────────────────────────────────────────────
#  CSS
# ──────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  .stApp { background: linear-gradient(135deg, #0b132b 0%, #1c2541 100%); }
  section[data-testid="stSidebar"] { background: #1e1b4b !important; }
  section[data-testid="stSidebar"] * { color: #7ecefd !important; }
  h1 { color: #ffd159 !important; text-align: center; }
  h2, h3 { color: #7ecefd !important; }
  .card {
    background: rgba(30,27,75,0.9);
    border: 1px solid #7ecefd44;
    border-radius: 14px;
    padding: 20px;
    margin: 10px 0;
  }
  .stButton > button {
    background: linear-gradient(90deg,#2e6bef,#7ecefd);
    color: white; border: none; border-radius: 10px;
    font-weight: bold; padding: .55rem 1.2rem;
  }
  .stButton > button:hover { opacity:.85; transform:scale(1.02); }
  label, .stTextInput label, .stSelectbox label,
  .stSlider label, .stNumberInput label { color:#7ecefd !important; font-weight:600; }
  .stTextInput > div > div > input,
  .stNumberInput > div > div > input,
  .stTextArea textarea {
    background:#0b132b !important; color:white !important;
    border:1px solid #7ecefd66 !important; border-radius:8px;
  }
  [data-testid="stMetricValue"] { color:#ffd159 !important; }
  [data-testid="stMetricLabel"] { color:#7ecefd !important; }
  hr { border-color:#7ecefd33 !important; }
  .stTabs [data-baseweb="tab"] { color:#7ecefd; background:#1e1b4b; border-radius:8px 8px 0 0; }
  .stTabs [aria-selected="true"] { background:#2e6bef !important; color:white !important; }
</style>
""", unsafe_allow_html=True)

# ──────────────────────────────────────────────────────────────────
#  Session State
# ──────────────────────────────────────────────────────────────────
DEFAULTS = {
    "template_bytes": None,
    "text_x": 50,
    "text_y": 60,
    "font_size": 72,
    "text_color": "#1a1a1a",
    "font_style": "Bold",
    "event_name": "Certificate of Participation",
    "event_topic": "",
    "event_date": str(date.today()),
    "event_venue": "",
    "organizer": "",
    "admin_auth": False,
    "admin_password": "admin123",
    "registered": {"Participant": [], "Teacher": [], "Speaker": [], "Management": []},
    "cert_log": [],
    "qr_data": None,
    "qr_url": "",
}
for k, v in DEFAULTS.items():
    if k not in st.session_state:
        st.session_state[k] = v

# ──────────────────────────────────────────────────────────────────
#  Font Map
# ──────────────────────────────────────────────────────────────────
FONT_MAP = {
    "Regular":      ["arial.ttf",   "DejaVuSans.ttf",          "FreeSans.ttf"],
    "Bold":         ["arialbd.ttf", "DejaVuSans-Bold.ttf",     "FreeSerifBold.ttf"],
    "Italic":       ["ariali.ttf",  "DejaVuSans-Oblique.ttf",  "FreeSansOblique.ttf"],
    "Bold Italic":  ["arialbi.ttf", "DejaVuSans-BoldOblique.ttf","FreeSansBoldOblique.ttf"],
    "Times":        ["times.ttf",   "DejaVuSerif.ttf",         "FreeSerif.ttf"],
    "Times Bold":   ["timesbd.ttf", "DejaVuSerif-Bold.ttf",    "FreeSerifBold.ttf"],
    "Courier":      ["cour.ttf",    "DejaVuSansMono.ttf",      "FreeMono.ttf"],
    "Courier Bold": ["courbd.ttf",  "DejaVuSansMono-Bold.ttf", "FreeMonoBold.ttf"],
}

def load_font(style: str, size: int) -> ImageFont.ImageFont:
    for fname in FONT_MAP.get(style, FONT_MAP["Bold"]):
        try:
            return ImageFont.truetype(fname, size)
        except Exception:
            continue
    return ImageFont.load_default()

def hex_to_rgba(h: str, alpha=255):
    h = h.lstrip("#")
    return (int(h[0:2],16), int(h[2:4],16), int(h[4:6],16), alpha)

def get_cfg() -> dict:
    return {
        "text_x":    st.session_state.text_x,
        "text_y":    st.session_state.text_y,
        "font_size": st.session_state.font_size,
        "text_color":st.session_state.text_color,
        "font_style":st.session_state.font_style,
    }

def get_event_info() -> dict:
    try:
        dt = datetime.strptime(st.session_state.event_date, "%Y-%m-%d")
        day = dt.strftime("%A")
    except Exception:
        day = ""
    return {
        "event_name": st.session_state.event_name,
        "topic":      st.session_state.event_topic,
        "event_date": st.session_state.event_date,
        "day":        day,
        "venue":      st.session_state.event_venue,
        "organizer":  st.session_state.organizer,
    }

# ──────────────────────────────────────────────────────────────────
#  Core Functions
# ──────────────────────────────────────────────────────────────────
def generate_certificate(name: str, template_bytes: bytes, c: dict) -> bytes:
    img = Image.open(io.BytesIO(template_bytes)).convert("RGBA")
    w, h = img.size
    font = load_font(c["font_style"], c["font_size"])
    px = int(w * c["text_x"] / 100)
    py = int(h * c["text_y"] / 100)
    layer = Image.new("RGBA", img.size, (255,255,255,0))
    draw  = ImageDraw.Draw(layer)
    bbox  = draw.textbbox((0,0), name, font=font)
    tw = bbox[2] - bbox[0]
    th = bbox[3] - bbox[1]
    draw.text((px - tw//2, py - th//2), name, font=font,
              fill=hex_to_rgba(c["text_color"]))
    final = Image.alpha_composite(img, layer).convert("RGB")
    buf = io.BytesIO()
    final.save(buf, format="PNG", dpi=(300,300))
    return buf.getvalue()

def png_to_pdf(png_bytes: bytes, name: str) -> bytes:
    buf    = io.BytesIO()
    pw, ph = landscape(A4)
    c      = pdf_canvas.Canvas(buf, pagesize=(pw, ph))
    img    = Image.open(io.BytesIO(png_bytes)).convert("RGB")
    iw, ih = img.size
    scale  = min(pw/iw, ph/ih)
    nw, nh = iw*scale, ih*scale
    x, y   = (pw-nw)/2, (ph-nh)/2
    tmp = io.BytesIO()
    img.save(tmp, format="PNG")
    tmp.seek(0)
    c.drawImage(ImageReader(tmp), x, y, nw, nh, mask="auto")
    c.setFont("Helvetica-Bold", 9)
    c.setFillColorRGB(.5,.5,.5)
    c.drawCentredString(pw/2, 16,
        f"{name} | {st.session_state.event_name} | "
        f"{datetime.now().strftime('%Y-%m-%d %H:%M')}")
    c.save()
    return buf.getvalue()

def make_qr(url: str) -> bytes:
    qr = qrcode.QRCode(
        version=1,
        error_correction=qrcode.constants.ERROR_CORRECT_H,
        box_size=10, border=4)
    qr.add_data(url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="#0b132b", back_color="white")
    buf = io.BytesIO()
    img.save(buf, format="PNG")
    return buf.getvalue()

def build_excel_report(event_info: dict, log: list) -> bytes:
    wb   = openpyxl.Workbook()
    hfil = PatternFill("solid", fgColor="1E1B4B")
    hfnt = Font(bold=True, color="FFFFFF", size=12)

    # ── Sheet 1: Event Summary ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Event Summary"
    ws1.merge_cells("A1:C1")
    t = ws1["A1"]
    t.value = f"🎓 {event_info.get('event_name','Event')} — Certificate Report"
    t.font  = Font(bold=True, color="FFD159", size=15)
    t.fill  = PatternFill("solid", fgColor="0B132B")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 36

    info_rows = [
        ("Event Name",   event_info.get("event_name","")),
        ("Topic",        event_info.get("topic","")),
        ("Date",         event_info.get("event_date","")),
        ("Day",          event_info.get("day","")),
        ("Venue",        event_info.get("venue","")),
        ("Organizer",    event_info.get("organizer","")),
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Total Certs",  str(len(log))),
    ]
    for r, (k, v) in enumerate(info_rows, 2):
        ws1[f"A{r}"] = k;  ws1[f"A{r}"].font = Font(bold=True, color="7ECEFD")
        ws1[f"B{r}"] = v;  ws1[f"B{r}"].font = Font(color="E0E0E0")
        ws1[f"A{r}"].fill = hfil
    ws1.column_dimensions["A"].width = 20
    ws1.column_dimensions["B"].width = 45

    # ── Sheet 2: Certificate Log ─────────────────────────────────
    ws2 = wb.create_sheet("Certificate Log")
    headers2 = ["#","Full Name","Department","Batch","Roll No","Category","Event","Date","Day","Time"]
    for ci, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=ci, value=h)
        cell.font = hfnt; cell.fill = hfil
        cell.alignment = Alignment(horizontal="center")
    for ri, rec in enumerate(log, 2):
        row_data = [
            ri-1,
            rec.get("name",""),
            rec.get("department",""),
            rec.get("batch",""),
            rec.get("roll_no",""),
            rec.get("category",""),
            rec.get("event",""),
            rec.get("date",""),
            rec.get("day",""),
            rec.get("time",""),
        ]
        for ci, val in enumerate(row_data, 1):
            c2 = ws2.cell(row=ri, column=ci, value=val)
            c2.font = Font(color="E0E0E0")
            c2.fill = PatternFill("solid", fgColor="0F1B35" if ri%2==0 else "1E1B4B")
            c2.alignment = Alignment(horizontal="center" if ci==1 else "left")
    for ci, w in enumerate([5,28,22,16,14,15,30,13,12,10], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w

    # ── Sheet 3: Category Summary ────────────────────────────────
    ws3 = wb.create_sheet("Category Summary")
    for ci, h in enumerate(["Category","Count","Names"], 1):
        cell = ws3.cell(row=1, column=ci, value=h)
        cell.font = hfnt; cell.fill = hfil
    categories: dict = {}
    for rec in log:
        cat = rec.get("category", "Other")
        categories.setdefault(cat, []).append(rec["name"])
    for ri, (cat, names) in enumerate(categories.items(), 2):
        ws3[f"A{ri}"] = cat;           ws3[f"A{ri}"].font = Font(bold=True, color="FFD159")
        ws3[f"B{ri}"] = len(names);    ws3[f"B{ri}"].font = Font(color="E0E0E0")
        # Build detailed name list with roll no
        detail_list = []
        for rec in log:
            if rec.get("category","") == cat:
                detail_list.append(f"{rec['name']} ({rec.get('roll_no','')})")
        display = ", ".join(detail_list) if detail_list else ", ".join(names)
        ws3[f"C{ri}"] = display; ws3[f"C{ri}"].font = Font(color="E0E0E0")
        for col in "ABC":
            ws3[f"{col}{ri}"].fill = hfil
    for col, w in [("A",20),("B",10),("C",70)]:
        ws3.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ══════════════════════════════════════════════════════════════════
#  ROUTING
# ══════════════════════════════════════════════════════════════════
qp   = st.query_params
page = qp.get("page", "admin")

# ══════════════════════════════════════════════════════════════════
#  STUDENT PAGE — NO LOGIN, just name input
# ══════════════════════════════════════════════════════════════════
if page == "cert":
    event   = qp.get("event",  "Certificate Event").replace("%20"," ")
    tx      = float(qp.get("tx", 50))
    ty      = float(qp.get("ty", 60))
    fs      = int(  qp.get("fs", 72))
    tc      = qp.get("tc", "#1a1a1a").replace("%23","#")
    fw      = qp.get("fw", "Bold").replace("%20"," ")
    cats_raw= qp.get("cats","Participant,Teacher,Speaker,Management")
    cat_opt = [c.replace("%20"," ") for c in cats_raw.split(",")]

    # Header
    st.markdown(f"""
    <div style="text-align:center;padding:30px 0 10px;">
      <h1 style="color:#ffd159;font-size:2.2rem;">🎓 {event}</h1>
      <p style="color:#7ecefd;font-size:1.1rem;">QR Certificate System</p>
    </div>
    """, unsafe_allow_html=True)
    st.markdown("---")

    if st.session_state.template_bytes is None:
        st.error("⚠️ Admin ne abhi template upload nahi kiya. Thodi der baad try karein.")
        st.stop()

    _, col, _ = st.columns([1,2,1])
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)

        category     = st.selectbox("🏷️ Category", cat_opt)
        name_input   = st.text_input("👤 Poora Naam / Full Name *",
                                     placeholder="e.g. Muhammad Ali Khan")
        dept_input   = st.text_input("🏫 Department *",
                                     placeholder="e.g. Computer Science")
        batch_input  = st.text_input("📅 Batch / Year *",
                                     placeholder="e.g. 2022-2026  or  Batch-7")
        rollno_input = st.text_input("🔢 Roll No *",
                                     placeholder="e.g. CS-2022-45")

        st.markdown("---")
        generate_clicked = st.button(
            "🎓 Certificate Generate Karein",
            use_container_width=True
        )

        if generate_clicked:
            name_clean   = name_input.strip()
            dept_clean   = dept_input.strip()
            batch_clean  = batch_input.strip()
            rollno_clean = rollno_input.strip()

            missing = []
            if not name_clean:   missing.append("Naam")
            if not dept_clean:   missing.append("Department")
            if not batch_clean:  missing.append("Batch")
            if not rollno_clean: missing.append("Roll No")

            if missing:
                st.error("❌ Yeh fields zaroori hain: " + ", ".join(missing))
            else:
                with st.spinner("🎨 Aapka certificate ban raha hai..."):
                    c_cfg = {"text_x":tx,"text_y":ty,"font_size":fs,
                             "text_color":tc,"font_style":fw}
                    png = generate_certificate(
                        name_clean, st.session_state.template_bytes, c_cfg)
                    pdf = png_to_pdf(png, name_clean)

                    now = datetime.now()
                    reg = st.session_state.registered
                    if category not in reg:
                        reg[category] = []
                    if name_clean not in reg[category]:
                        reg[category].append(name_clean)

                    st.session_state.cert_log.append({
                        "name":       name_clean,
                        "department": dept_clean,
                        "batch":      batch_clean,
                        "roll_no":    rollno_clean,
                        "category":   category,
                        "event":      event,
                        "date":       now.strftime("%Y-%m-%d"),
                        "day":        now.strftime("%A"),
                        "time":       now.strftime("%H:%M:%S"),
                    })

                st.success(f"✅ Certificate tayar hai — **{name_clean}**!")
                st.image(png, use_container_width=True)

                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "⬇️ PNG Download", png,
                        file_name=f"Certificate_{name_clean}.png",
                        mime="image/png", use_container_width=True)
                with c2:
                    st.download_button(
                        "⬇️ PDF Download", pdf,
                        file_name=f"Certificate_{name_clean}.pdf",
                        mime="application/pdf", use_container_width=True)
                st.balloons()

        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown("---")
    st.markdown(
        '<p style="text-align:center;color:#7ecefd;font-size:.85rem;">'
        '🔒 QR Certificate System | Developed by Abdul Samad — SBBU Nawabshah</p>',
        unsafe_allow_html=True)
    st.stop()


# ══════════════════════════════════════════════════════════════════
#  ADMIN PAGE
# ══════════════════════════════════════════════════════════════════
st.markdown("# 🎓 QR Certificate Generator Pro v2.0")
st.markdown(
    '<p style="text-align:center;color:#7ecefd;">'
    'Abdul Samad | Shaheed Benazir Bhutto University Nawabshah</p>',
    unsafe_allow_html=True)
st.markdown("---")

# ── Auth ─────────────────────────────────────────────────────────
if not st.session_state.admin_auth:
    _, col, _ = st.columns([1,2,1])
    with col:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🔐 Admin Login")
        pwd = st.text_input("Password", type="password")
        if st.button("Login", use_container_width=True):
            if pwd == st.session_state.admin_password:
                st.session_state.admin_auth = True
                st.rerun()
            else:
                st.error("❌ Galat password!")
        st.caption("Default: admin123")
        st.markdown('</div>', unsafe_allow_html=True)
    st.stop()

# ── Sidebar ───────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Text Settings")
    st.session_state.font_size   = st.slider("Font Size", 20, 200, st.session_state.font_size)
    st.session_state.text_x     = st.slider("Horizontal % (←→)", 0, 100, st.session_state.text_x)
    st.session_state.text_y     = st.slider("Vertical %   (↑↓)", 0, 100, st.session_state.text_y)
    st.session_state.text_color = st.color_picker("Text Color", st.session_state.text_color)
    st.session_state.font_style = st.selectbox(
        "Font Style", list(FONT_MAP.keys()),
        index=list(FONT_MAP.keys()).index(st.session_state.font_style))
    st.markdown("---")
    st.markdown("## 📋 Event Info")
    st.session_state.event_name  = st.text_input("Event Name",        st.session_state.event_name)
    st.session_state.event_topic = st.text_input("Topic",             st.session_state.event_topic)
    st.session_state.event_date  = st.text_input("Date (YYYY-MM-DD)", st.session_state.event_date)
    st.session_state.event_venue = st.text_input("Venue",             st.session_state.event_venue)
    st.session_state.organizer   = st.text_input("Organizer",         st.session_state.organizer)
    st.markdown("---")
    with st.expander("🔑 Change Password"):
        np_ = st.text_input("New Password", type="password", key="np_")
        if st.button("Update", key="upd_pwd") and np_:
            st.session_state.admin_password = np_
            st.success("✅ Updated!")
    if st.button("🚪 Logout"):
        st.session_state.admin_auth = False
        st.rerun()

# ── Tabs ──────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📁 Setup & QR",
    "👁️ Preview & Edit",
    "📊 Bulk Generate",
    "📈 Analytics & Report",
    "☁️ GitHub Deploy Guide",
])

# ════════════════════════════════════════════════════
#  TAB 1 — Setup & QR
# ════════════════════════════════════════════════════
with tab1:
    cl, cr = st.columns(2)

    with cl:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📁 Step 1: Template Upload")
        upl = st.file_uploader(
            "Certificate template (.png/.jpg)",
            type=["png","jpg","jpeg"])
        if upl:
            st.session_state.template_bytes = upl.read()
            img_tmp = Image.open(io.BytesIO(st.session_state.template_bytes))
            st.success(f"✅ {upl.name} — {img_tmp.width}×{img_tmp.height}px")
            st.image(st.session_state.template_bytes, use_container_width=True)
        elif st.session_state.template_bytes:
            st.image(st.session_state.template_bytes,
                     caption="Current Template", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📄 Step 2: Names List Upload")
        names_upl = st.file_uploader(
            "Upload .txt (one name per line)", type=["txt"])
        cat_for_upload = st.selectbox(
            "Kin ke liye hain yeh names?",
            ["Participant","Teacher","Speaker","Management"])
        if names_upl:
            names_raw  = names_upl.read().decode("utf-8")
            names_list = [n.strip() for n in names_raw.splitlines() if n.strip()]
            existing   = st.session_state.registered.get(cat_for_upload, [])
            merged     = list(dict.fromkeys(existing + names_list))
            st.session_state.registered[cat_for_upload] = merged
            st.success(f"✅ {len(names_list)} names loaded → {cat_for_upload}")
        st.markdown('</div>', unsafe_allow_html=True)

    with cr:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 🔗 Step 3: Generate QR Code")

        app_url = st.text_input(
            "Your Deployed App URL",
            value="https://your-app.streamlit.app",
            help="GitHub deploy ke baad URL yahan paste karein")

        cats_input = st.text_input(
            "Categories (comma separated)",
            value="Participant,Teacher,Speaker,Management")

        if st.button("🔳 Generate QR Code", use_container_width=True):
            if not st.session_state.template_bytes:
                st.warning("⚠️ Pehle template upload karein!")
            else:
                c = get_cfg()
                tc_enc  = c['text_color'].replace('#','%23')
                ev_enc  = st.session_state.event_name.replace(' ','%20')
                fw_enc  = c['font_style'].replace(' ','%20')
                cats_enc= cats_input.replace(' ','%20')
                url = (f"{app_url.rstrip('/')}/?page=cert"
                       f"&event={ev_enc}"
                       f"&tx={c['text_x']}&ty={c['text_y']}"
                       f"&fs={c['font_size']}&tc={tc_enc}"
                       f"&fw={fw_enc}&cats={cats_enc}")
                st.session_state.qr_url  = url
                st.session_state.qr_data = make_qr(url)

        if st.session_state.qr_data:
            st.image(st.session_state.qr_data, width=240)
            st.code(st.session_state.qr_url, language=None)
            st.download_button(
                "⬇️ Download QR PNG",
                data=st.session_state.qr_data,
                file_name="event_qr.png", mime="image/png",
                use_container_width=True)
            st.info("📌 Print karo → Event mein lagao → Students scan karein!")
        st.markdown('</div>', unsafe_allow_html=True)

        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("### 📱 Student Flow (QR Scan ke Baad)")
        st.markdown("""
1. 📱 QR scan → browser khulta hai (koi login nahi!)  
2. Category choose karo (Participant / Teacher…)  
3. ✏️ Naam likho  
4. 🎓 Certificate instantly milta hai  
5. ⬇️ PNG ya PDF download karo  
6. ✅ Naam automatically admin ke list mein save  
        """)
        st.markdown('</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════
#  TAB 2 — Preview & Edit
# ════════════════════════════════════════════════════
with tab2:
    if not st.session_state.template_bytes:
        st.warning("⚠️ Pehle Tab 1 mein template upload karein.")
    else:
        # ── Single preview ─────────────────────────
        st.markdown("### 🔍 Live Single Preview")
        st.info("💡 Sidebar se font size, position, color adjust karein — preview instantly update hoga!")

        prev_name = st.text_input(
            "Preview ke liye naam likhein:",
            value="Muhammad Ali Khan", key="prev_name")

        png_prev = generate_certificate(
            prev_name, st.session_state.template_bytes, get_cfg())
        st.image(png_prev, use_container_width=True,
                 caption=(f"Preview: {prev_name} | "
                          f"Size: {st.session_state.font_size} | "
                          f"Pos: ({st.session_state.text_x}%, {st.session_state.text_y}%) | "
                          f"Color: {st.session_state.text_color} | "
                          f"Font: {st.session_state.font_style}"))

        ca, cb = st.columns(2)
        with ca:
            st.download_button(
                "⬇️ PNG Download", png_prev,
                file_name=f"Preview_{prev_name}.png",
                mime="image/png", use_container_width=True)
        with cb:
            st.download_button(
                "⬇️ PDF Download",
                png_to_pdf(png_prev, prev_name),
                file_name=f"Preview_{prev_name}.pdf",
                mime="application/pdf", use_container_width=True)

        st.markdown("---")

        # ── All-names preview ───────────────────────
        st.markdown("### 👁️ Preview All Names (Sabke Certificates Dekho)")
        st.markdown("Scroll karke har naam ka certificate check karo — koi galti nahi rahegi!")

        all_names = [(n, cat)
                     for cat, nms in st.session_state.registered.items()
                     for n in nms]

        if not all_names:
            st.info("Koi naam abhi list mein nahi hai. Tab 1 se .txt upload karein ya QR scan karwain.")
        else:
            max_prev = min(len(all_names), 30)
            show_n   = st.slider(
                "Kitne certificates preview karein?",
                1, max_prev, min(6, max_prev))

            cols_per_row = 3
            subset = all_names[:show_n]
            for i in range(0, len(subset), cols_per_row):
                row_items = subset[i:i+cols_per_row]
                cs = st.columns(cols_per_row)
                for ci, (nm, cat) in enumerate(row_items):
                    with cs[ci]:
                        pv = generate_certificate(
                            nm, st.session_state.template_bytes, get_cfg())
                        st.image(pv, caption=f"[{cat}] {nm}",
                                 use_container_width=True)
                        st.download_button(
                            f"⬇️ {nm[:18]}",
                            data=pv,
                            file_name=f"{nm}.png",
                            mime="image/png",
                            key=f"dl_{nm}_{i}_{ci}")


# ════════════════════════════════════════════════════
#  TAB 3 — Bulk Generate
# ════════════════════════════════════════════════════
with tab3:
    st.markdown("### 📊 Bulk Certificate Generation")

    if not st.session_state.template_bytes:
        st.warning("⚠️ Pehle Tab 1 mein template upload karein.")
    else:
        # ── Names Editor ─────────────────────────────
        st.markdown("#### ✏️ Names Review & Edit (Category wise)")

        for cat in list(st.session_state.registered.keys()):
            n_count = len(st.session_state.registered[cat])
            with st.expander(f"📂 {cat} — {n_count} names"):
                raw_text = st.text_area(
                    f"Names — ek naam per line",
                    value="\n".join(st.session_state.registered[cat]),
                    height=160,
                    key=f"edit_{cat}")
                if st.button(f"💾 Save {cat} List", key=f"save_{cat}"):
                    updated = [n.strip() for n in raw_text.splitlines() if n.strip()]
                    st.session_state.registered[cat] = updated
                    st.success(f"✅ {cat} list updated — {len(updated)} names")
                    st.rerun()

        st.markdown("---")

        # ── Stats ────────────────────────────────────
        all_flat = [(n, cat)
                    for cat, nms in st.session_state.registered.items()
                    for n in nms]

        mcols = st.columns(len(st.session_state.registered)+1)
        mcols[0].metric("Total", len(all_flat))
        for i, (cat, nms) in enumerate(st.session_state.registered.items()):
            mcols[i+1].metric(cat, len(nms))

        st.markdown("---")
        st.markdown("#### 🚀 Sab Generate Karo + Download Karo")

        if not all_flat:
            st.info("Koi naam nahi hai. Upar edit karein ya Tab 1 se upload karein.")
        else:
            if st.button(
                f"🚀 Generate All {len(all_flat)} Certificates (ZIP)",
                use_container_width=True):

                prog   = st.progress(0)
                status = st.empty()
                records= []
                buf_zip= io.BytesIO()

                with zipfile.ZipFile(buf_zip, "w", zipfile.ZIP_DEFLATED) as zf:
                    for i, (nm, cat) in enumerate(all_flat):
                        status.markdown(f"⏳ **{nm}** [{cat}] ({i+1}/{len(all_flat)})")
                        png = generate_certificate(
                            nm, st.session_state.template_bytes, get_cfg())
                        zf.writestr(f"{cat}/{nm}.png", png)
                        now = datetime.now()
                        rec = {
                            "name":nm, "category":cat,
                            "event":st.session_state.event_name,
                            "date":now.strftime("%Y-%m-%d"),
                            "time":now.strftime("%H:%M:%S")
                        }
                        records.append(rec)
                        # Add to log if not already there
                        existing_names = [r["name"] for r in st.session_state.cert_log]
                        if nm not in existing_names:
                            st.session_state.cert_log.append(rec)
                        prog.progress((i+1)/len(all_flat))

                status.success(f"✅ {len(all_flat)} certificates ready!")

                c1, c2 = st.columns(2)
                with c1:
                    st.download_button(
                        "⬇️ Download All (ZIP)",
                        data=buf_zip.getvalue(),
                        file_name=f"{st.session_state.event_name}_Certificates.zip",
                        mime="application/zip",
                        use_container_width=True)
                with c2:
                    excel = build_excel_report(get_event_info(), records)
                    st.download_button(
                        "📊 Download Excel Report",
                        data=excel,
                        file_name=f"{st.session_state.event_name}_Report.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True)


# ════════════════════════════════════════════════════
#  TAB 4 — Analytics & Report
# ════════════════════════════════════════════════════
with tab4:
    st.markdown("### 📈 Event Analytics & Certificate Report")

    reg   = st.session_state.registered
    total = sum(len(v) for v in reg.values())
    log   = st.session_state.cert_log

    # Metrics
    m_cols = st.columns(len(reg)+2)
    m_cols[0].metric("Total Registered", total)
    m_cols[1].metric("QR Scans / Certs",  len(log))
    for i, (cat, nms) in enumerate(reg.items()):
        m_cols[i+2].metric(cat, len(nms))

    st.markdown("---")

    # Event Info Summary
    ei = get_event_info()
    with st.expander("📋 Event Info Summary"):
        for k, v in ei.items():
            st.markdown(f"**{k.replace('_',' ').title()}:** {v}")

    st.markdown("---")
    st.markdown("#### 📋 Live Registration Log (QR Scan se aaye names)")

    if log:
        # Show as formatted table with friendly column names
        import pandas as pd
        df = pd.DataFrame(log)
        col_rename = {
            "name":"Full Name","department":"Department","batch":"Batch",
            "roll_no":"Roll No","category":"Category","event":"Event",
            "date":"Date","day":"Day","time":"Time"
        }
        df = df.rename(columns={k:v for k,v in col_rename.items() if k in df.columns})
        st.dataframe(df, use_container_width=True)

        c1, c2, c3 = st.columns(3)
        with c1:
            excel = build_excel_report(ei, log)
            st.download_button(
                "📊 Full Excel Report",
                data=excel,
                file_name=f"{st.session_state.event_name}_Full_Report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        with c2:
            names_txt = "\n".join(
                f"[{r['category']}] {r['name']}" for r in log)
            st.download_button(
                "📄 Names TXT Export",
                data=names_txt.encode(),
                file_name="registered_names.txt",
                mime="text/plain",
                use_container_width=True)
        with c3:
            if st.button("🗑️ Clear Log", use_container_width=True):
                st.session_state.cert_log = []
                st.rerun()
    else:
        st.info("Abhi koi QR scan nahi hua. Students scan karein to yahan naam aayenge.")

    # Registered names summary per category
    st.markdown("---")
    st.markdown("#### 📂 Registered Names (Category wise)")
    for cat, nms in reg.items():
        if nms:
            st.markdown(f"**{cat}** ({len(nms)}): {', '.join(nms)}")


# ════════════════════════════════════════════════════
#  TAB 5 — GitHub Deploy Guide
# ════════════════════════════════════════════════════
with tab5:
    st.markdown("### ☁️ Free Hosting — GitHub + Streamlit Cloud")
    st.markdown("""
<div class="card">

## 🚀 Ek Baar Setup — Hamesha Free Online!

### ✅ Step 1 — GitHub Account Banao
1. [github.com](https://github.com) par jao → **Sign Up** (free)
2. **New Repository** click karo
3. Name: `qr-certificate-generator`
4. **Public** rakho → **Create Repository**

---

### ✅ Step 2 — Files Upload Karo
Repository page par **"uploading an existing file"** click karo aur yeh 2 files upload karo:
```
app.py
requirements.txt
```
Ya PowerShell mein Git use karo:
```bash
cd d:/Avalon.AI
git init
git add app.py requirements.txt
git commit -m "first commit"
git branch -M main
git remote add origin https://github.com/YOUR_USERNAME/qr-certificate-generator.git
git push -u origin main
```

---

### ✅ Step 3 — Streamlit Cloud Deploy
1. [share.streamlit.io](https://share.streamlit.io) par jao
2. **GitHub se Login** karo
3. **"New app"** click karo
4. Repository: `qr-certificate-generator`
5. Branch: `main`
6. Main file path: `app.py`
7. **"Deploy!"** click karo

⏳ **2-3 minute mein live!**

---

### ✅ Step 4 — QR Generate Karo
Deploy hone ke baad URL milega:
```
https://YOUR_USERNAME-qr-certificate-generator-app-XXXXX.streamlit.app
```

Yeh URL **Tab 1 → "Your Deployed App URL"** mein paste karo → **QR Generate** → **Print** karo ✅

---

### ⚠️ Important Note (Template ke Baare Mein)
Streamlit Cloud **RAM-based session** use karta hai.  
Matlab har baar admin ko template upload karna hoga jab app restart ho.

**Permanent solution ke liye (Advanced):**
- Template ko GitHub repo mein rakh do (default template)
- Ya Google Drive API use karo

---

### ✅ Cost & Benefits Summary

| Feature | Detail |
|---------|--------|
| 💰 Cost | Bilkul Free |
| 🌐 Access | Kahi se bhi |
| 👥 Concurrent Users | 100+ |
| 🔄 Update | Git push → auto update |
| ⏰ Uptime | 24/7 |
| 📱 Mobile | Fully supported |

</div>
    """, unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────
st.markdown("---")
st.markdown(
    '<p style="text-align:center;color:#7ecefd;font-size:.9rem;">'
    '© QR Certificate Generator Pro v2.0 | Abdul Samad | SBBU Nawabshah</p>',
    unsafe_allow_html=True)
